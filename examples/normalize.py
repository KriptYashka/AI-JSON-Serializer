import asyncio
import sys
from pathlib import Path

from openpyxl import load_workbook

from src.pipeline import run_pipeline
from src.schema import Normalizer, load_schema


def _read_headers(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(values_only=True))]
    wb.close()
    return headers


async def main() -> None:
    if len(sys.argv) < 2:
        print("Использование: python -m examples.normalize <input.xlsx> [--output <out.xlsx>]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.is_file():
        print(f"Файл не найден: {input_path}")
        sys.exit(1)

    output_arg: str | None = None
    for i, a in enumerate(sys.argv[2:]):
        if a == "--output" and i + 1 < len(sys.argv[2:]):
            output_arg = sys.argv[2:][i + 1]

    schema = load_schema()
    headers = _read_headers(input_path)

    print(f"Заголовки: {headers}")
    print("Цепочка: дешёвая модель → дорогая проверка\n")

    mapping_rules = await run_pipeline(headers, schema)

    print("\nИтоговый маппинг:")
    for r in mapping_rules:
        print(f"  {r.from_} → {r.to}")

    output_path = Path(output_arg or f"normalized_{input_path.name}")
    normalizer = Normalizer(schema, mapping_rules)
    count = normalizer.normalize(input_path, output_path)
    print(f"\nГотово: {count} строк → {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
