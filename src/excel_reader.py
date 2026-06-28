import re
from pathlib import Path
from typing import Any, Generator


def _reader_openpyxl(path: str | Path) -> Generator[list[Any], None, None]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        yield [str(c).strip() if c is not None else "" for c in row]
    wb.close()


def _reader_xlrd(path: str | Path) -> Generator[list[Any], None, None]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    sheet = wb.sheet_by_index(0)
    for r in range(sheet.nrows):
        yield [str(c).strip() if c != "" else "" for c in sheet.row_values(r)]


def read_excel(path: str | Path) -> Generator[list[Any], None, None]:
    path = Path(path)
    if path.suffix.lower() == ".xls":
        yield from _reader_xlrd(path)
    else:
        yield from _reader_openpyxl(path)


def read_headers(path: str | Path, header_row: int = 0) -> list[str]:
    for i, row in enumerate(read_excel(path)):
        if i == header_row:
            return row
    return []


def detect_header_row(path: str | Path, max_rows: int = 10, min_cols: int = 2) -> int:
    rows: list[list[str]] = []
    for i, row in enumerate(read_excel(path)):
        if i >= max_rows:
            break
        rows.append(row)

    best_score = -1.0
    best_row = 0
    second_best = -1.0

    for i, row in enumerate(rows):
        non_empty = [c for c in row if c]
        if len(non_empty) < min_cols:
            continue

        score = 0.0
        score += len(non_empty) * 2.0
        unique = set(c.lower() for c in non_empty)
        score += len(unique) * 1.5
        alpha_count = sum(1 for c in non_empty if any(ch.isalpha() for ch in c))
        score += alpha_count * 1.0
        number_count = sum(1 for c in non_empty if _looks_like_number(c))
        score -= number_count * 3.0
        long_text = sum(1 for c in non_empty if len(c) > 60)
        score -= long_text * 4.0
        score /= max(len(non_empty), 1)

        if score > best_score:
            second_best = best_score
            best_score = score
            best_row = i
        elif score > second_best:
            second_best = score

    return best_row


def _looks_like_number(s: str) -> bool:
    return bool(re.match(r"^[\d\s.,₽$€%+-]+$", s.strip()))


async def ai_detect_header_row(path: str | Path, profile: str = "cheap") -> int:
    from src.client import OpenRouterClient
    from src.config import settings
    from src.models import ChatRequest, Message
    from src.schema.loader import load_prompt

    rows: list[list[str]] = []
    for i, row in enumerate(read_excel(path)):
        if i >= 30:
            break
        rows.append(row)

    lines = "\n".join(
        f"[{i}] {' | '.join(c if c else '(empty)' for c in row)}"
        for i, row in enumerate(rows)
    )

    prompt_tpl = load_prompt("detect_headers.txt")
    prompt = prompt_tpl.format(rows=lines)

    model = settings.profiles[profile].models[0]
    messages = [Message(role="system", content=prompt)]
    request = ChatRequest(model=model, messages=messages, temperature=0.0, max_tokens=50)

    async with OpenRouterClient() as client:
        response = await client.chat_completion(request)
        content = response.choices[0].message.content.strip()

    content = content.splitlines()[0].strip()
    import re as _re
    nums = _re.findall(r"-?\d+", content)
    return int(nums[0]) if nums else 0
